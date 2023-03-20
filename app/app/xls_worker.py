import random
from enum import Enum

import openpyxl
import pandas as pd
from fastapi import UploadFile
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame
from roman import toRoman
from fastapi.responses import FileResponse


class Style(Enum):
    FONT = Font(bold=True)
    FILL = PatternFill("solid", fgColor="00C0C0C0")
    DATA_FILL = PatternFill("solid", fgColor="00CCFFCC")


def get_xls_data(file: UploadFile):
    if file.filename.split('.')[-1] == 'csv':
        csv = pd.read_csv(file.file)

        try:
            if type(csv.values[0][2]) == float:
                return zip(csv[csv.keys()[0]], csv[csv.keys()[1]], csv[csv.keys()[2]]), True
        except:
            pass

        return zip(csv[csv.keys()[0]], csv[csv.keys()[1]]), False

    else:
        xls = pd.ExcelFile(file.file)
        sheet: DataFrame = xls.parse(0)

        try:
            if type(sheet.values[0][2]) == float:
                return zip(sheet[sheet.keys()[0]], sheet[sheet.keys()[1]], sheet[sheet.keys()[2]]), True
        except:
            pass

        return zip(sheet[sheet.keys()[0]], sheet[sheet.keys()[1]]), False


def initialize_workbook(titles):
    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40

    insert_data(ws, titles, 1, True)

    return wb, ws


def insert_data(ws, data, row_counter, is_header=False):
    if not is_header:
        for column, value in enumerate(data):
            cell = ws.cell(row=row_counter, column=column + 1, value=value)
            if row_counter % 2 == 1:
                cell.fill = Style.DATA_FILL.value
    else:
        for column, value in enumerate(data):
            cell = ws.cell(row=row_counter, column=column + 1, value=value)
            cell.font = Style.FONT.value
            cell.fill = Style.FILL.value


def month_day_report(dataset, year):
    savefile = f'files/xls{random.randint(0, 100)}.xls'
    wb, ws = initialize_workbook(['Date', 'Year', 'Days Hashrate (EH)'])
    row_counter = 2

    date_hash_sum = dataset.groupby('date').hash.sum()

    dates = []

    for day, hash_rate, average, month_name, date in dataset[['day', 'hash', 'average', 'month_name', 'date']].values:
        if date not in dates:
            insert_data(ws, [f'{month_name} {day}, {year}', year, f"{round(date_hash_sum.get(date), 2):_.2f}".replace("_", " ")], row_counter)
            row_counter += 1
            dates.append(date)

    insert_data(ws, ['Totals:', year, f"{round(dataset.hash.sum(), 2):_.2f}".replace("_", " ")], row_counter, True)

    wb.save(savefile)

    return FileResponse(savefile)


def year_quarter_month_report(dataset, quarter_groups, months_sum, quarter_sum, year):
    savefile = f'files/xls{random.randint(0, 100)}.xls'
    wb, ws = initialize_workbook(['Month', 'Year', 'Month hashrate (EH)'])
    row_counter = 2

    for quarter in quarter_groups:
        for month_name in dataset.loc[dataset.quarter == quarter[0]].month_name.unique():
            insert_data(ws, [month_name, year, f"{round(months_sum.get(month_name), 2):_.2f}".replace("_", " ")], row_counter)
            row_counter += 1

        insert_data(ws, [f'{toRoman(quarter[0])} Quarter', year, f"{round(quarter_sum.get(quarter[0]), 2):_.2f}".replace("_", " ")], row_counter, True)
        row_counter += 1

    insert_data(ws, ['Totals:', year, f"{round(dataset.hash.sum(), 2):_.2f}".replace("_", " ")], row_counter, True)

    wb.save(savefile)

    return FileResponse(savefile)


def year_quarter_report(dataset, quarters_sum, year):
    savefile = f'files/xls{random.randint(0, 100)}.xls'
    wb, ws = initialize_workbook(['Quarter', 'Year', 'Quarter hashrate (EH)'])
    row_counter = 2

    for quarter_pk, quarter_sum in quarters_sum.items():
        insert_data(ws, [f'{toRoman(quarter_pk)} Quarter', year, f"{round(quarter_sum, 2):_.2f}".replace("_", " ")], row_counter)
        row_counter += 1

    insert_data(ws, ['Totals:', year, f"{round(dataset.hash.sum(), 2):_.2f}".replace("_", " ")], row_counter, True)

    wb.save(savefile)

    return FileResponse(savefile)


def year_quarter_month_day_report(dataset, quarter_groups, year, months_sum, quarter_sum, months_sum_average, quarter_sum_average):
    savefile = f'files/xls{random.randint(0, 100)}.xls'
    wb, ws = initialize_workbook(['Day/Months/Quarters', 'Average Hashrate (PH/s)', 'Day/Months/Quarters Hashrate (EH)'])
    row_counter = 2

    date_hash_sum = dataset.groupby('date').hash.sum()
    date_average_sum = dataset.groupby('date').average.sum()

    dates = []

    for quarter in quarter_groups:
        for month_name in dataset.loc[dataset.quarter == quarter[0]].month_name.unique():
            for day_ds in dataset.loc[dataset.month_name == month_name][['day', 'hash', 'month_name', 'average', 'date']].values:
                if day_ds[4] not in dates:
                    insert_data(ws, [f'{month_name} {day_ds[0]}, {year}', f"{round(date_average_sum.get(day_ds[4]), 3):_.3f}".replace("_", " "), f"{round(date_hash_sum.get(day_ds[4]), 2):_.2f}".replace("_", " ")], row_counter)
                    row_counter += 1
                    dates.append(day_ds[4])

            insert_data(ws, [f'{month_name} Total', f"{round(months_sum_average.get(month_name), 3):_.3f}".replace("_", " "), f"{round(months_sum.get(month_name), 2):_.2f}".replace("_", " ")], row_counter, True)
            row_counter += 1

        insert_data(ws, [f'{toRoman(quarter[0])} Quarter', f"{round(quarter_sum_average.get(quarter[0]), 3):_.3f}".replace("_", " "), f"{round(quarter_sum.get(quarter[0]), 2):_.2f}".replace("_", " ")], row_counter, True)
        row_counter += 1

    insert_data(ws, ['Totals:', f"{round(dataset.average.sum(), 3):_.3f}".replace("_", " "), f"{round(dataset.hash.sum(), 2):_.2f}".replace("_", " ")], row_counter, True)

    wb.save(savefile)

    return FileResponse(savefile)


def quarter_month_report(dataset, month_sums, month_names, year, quarter):
    savefile = f'files/xls{random.randint(0, 100)}.xls'
    wb, ws = initialize_workbook(['Month', 'Year', 'Months/Quarterly Hashrate (EH)'])
    row_counter = 2

    for (month_pk, month_sum), month_name in zip(month_sums.items(), month_names):
        insert_data(ws, [month_name, year, f"{round(month_sum, 2):_.2f}".replace("_", " ")], row_counter)
        row_counter += 1

    insert_data(ws, [f'Totals {toRoman(quarter)} Quarter:', year, f"{round(dataset.hash.sum(), 2):_.2f}".replace("_", " ")], row_counter, True)

    wb.save(savefile)

    return FileResponse(savefile)


def quarter_month_day_report(dataset, months_sum, year, quarter):
    savefile = f'files/xls{random.randint(0, 100)}.xls'
    wb, ws = initialize_workbook(['Date', 'Year', 'Days/Months Hashrate (EH)'])
    row_counter = 2

    date_hash_sum = dataset.groupby('date').hash.sum()

    dates = []

    for month_name in dataset.month_name.unique():
        for day_ds in dataset.loc[dataset.month_name == month_name][['day', 'hash', 'month_name', 'date']].values:
            if day_ds[3] not in dates:
                insert_data(ws, [f'{month_name} {day_ds[0]}, {year}', year, f"{round(date_hash_sum.get(day_ds[3]), 2):_.2f}".replace("_", " ")], row_counter)
                row_counter += 1
                dates.append(day_ds[3])

        insert_data(ws, [f'{month_name} Total', year, f"{round(months_sum.get(month_name), 2):_.2f}".replace("_", " ")], row_counter, True)
        row_counter += 1

    insert_data(ws, [f'Totals {toRoman(quarter)} Quarter:', year, f"{round(dataset.hash.sum(), 2):_.2f}".replace("_", " ")], row_counter, True)

    wb.save(savefile)

    return FileResponse(savefile)

