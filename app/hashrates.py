import numpy
import openpyxl
import pandas as pd
from fastapi import HTTPException
from openpyxl.worksheet.worksheet import Worksheet
from pandas import Series
from pandas.core.groupby import DataFrameGroupBy
from sqlalchemy import extract
from sqlalchemy.orm import Session

from models.db import hashrates as db_hashrates, users as db_users
from models.dto import hashrates as dto_hashrates
from app import xls_worker
from app.db import engine


def create_hashrate(db: Session, hashrate: dto_hashrates.HashrateBase):
    db_hash = db_hashrates.Hashrate(date=hashrate.date, average=hashrate.average, hash=hashrate.hash, company_id=hashrate.company_id)
    db.add(db_hash)
    db.commit()
    db.refresh(db_hash)
    return db_hash


def get_hashrate_by_company_id(db: Session, company_id: int, access_level, from_user_id):
    if access_level == 1:
        return db.query(db_hashrates.Hashrate).filter(db_hashrates.Hashrate.company_id == company_id).all()
    else:
        db.query(db_users.UserCompany).filter()
        return db.query(db_hashrates.Hashrate).join(db_users.UserCompany).filter(db_users.UserCompany.company_id == company_id).all()


def get_data_from_file(file, db, company_id, user_id):
    data = xls_worker.get_xls_data(file)
    hashrate_list = []
    for date, hashrate in data:
        db_hashrate = db.query(db_hashrates.Hashrate).filter(db_hashrates.Hashrate.date == date).filter(db_hashrates.Hashrate.company_id == company_id).first()
        if not db_hashrate:
            to_db_hashrate = db_hashrates.Hashrate(date=date, average=round(hashrate/86400*1000, 2), hash=hashrate, company_id=company_id, user_id=user_id)
            to_output_hashrate = to_db_hashrate.__dict__
            to_output_hashrate['status'] = 'new'
            hashrate_list.append(to_output_hashrate)
            db.add(to_db_hashrate)
        else:
            db_hashrate.average = round(hashrate/86400*1000, 2)
            db_hashrate.hash = hashrate
            db_hashrate.user_id = user_id
            to_output_hashrate = db_hashrate.__dict__
            to_output_hashrate['status'] = 'updated'
            hashrate_list.append(to_output_hashrate)

        db.commit()
    return hashrate_list


def get_hashrate_by_user_id(db, user_id):
    companies_id = db.query(db_users.UserCompany.company_id).filter(db_users.UserCompany.user_id == user_id)
    hashrates = []
    for company_id in companies_id:
        hashrates.extend(db.query(db_hashrates.Hashrate).filter(db_hashrates.Hashrate.company_id == company_id[0]))
    return hashrates


def get_all_hashrates(db):
    return db.query(db_hashrates.Hashrate).all()


def check_report_type(date_type: str):
    if date_type not in ['year_by_quarters', 'year_by_months', 'year_by_days',
                         'quarter_by_months', 'quarter_by_days', 'month_by_days']:
        raise HTTPException(status_code=404, detail="Date_type format failed")


def get_report(db: Session, file_format, company_id, from_date, to_date, auth, year):
    statement = db.query(db_hashrates.Hashrate).filter(extract('year', db_hashrates.Hashrate.date) == year)
    if company_id:
        statement = statement.filter(db_hashrates.Hashrate.company_id == company_id)
    if from_date:
        statement = statement.filter(db_hashrates.Hashrate.date >= from_date)
    if to_date:
        statement = statement.filter(db_hashrates.Hashrate.date <= to_date)

    statement = statement.statement

    dataset = pd.read_sql(statement, engine)
    dataset['date'] = pd.to_datetime(dataset.date, format='%Y-%m-%d')

    dataset['day'] = dataset.date.dt.day
    dataset['month']: Series = dataset.date.dt.month
    dataset['quarter'] = dataset.date.dt.quarter

    months_sum: Series = dataset.groupby('month').hash.sum()
    quarters_sum: Series = dataset.groupby('quarter').hash.sum()
    year_sum = quarters_sum.sum()

    if file_format == 'excel':
        wb = openpyxl.Workbook()
        ws: Worksheet = wb.active
        row_counter = 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15

        ws.cell(row=row_counter, column=1, value=f'{year}-year')
        ws.cell(row=row_counter, column=2, value=year_sum)
        row_counter += 1
        for quarter in dataset.quarter.unique():
            ws.cell(row=row_counter, column=1, value=f'{quarter}-quarter')
            ws.cell(row=row_counter, column=2, value=quarters_sum.get(quarter))
            row_counter += 1
            for month in dataset.loc[dataset.quarter == quarter].month.unique():
                ws.cell(row=row_counter, column=1, value=f'{month}-month')
                ws.cell(row=row_counter, column=2, value=months_sum.get(month))
                row_counter += 1
                for day, day_value in dataset.loc[dataset.month == month][['day', 'hash']].values:
                    ws.cell(row=row_counter, column=1, value=f'{int(day)}-day')
                    ws.cell(row=row_counter, column=2, value=day_value)
                    row_counter += 1

        wb.save("sample.xlsx")
        return
    elif file_format == 'pdf':
        pass
    else:
        report = {'year': {'total': year_sum}}
        for quarter in dataset.quarter.unique():
            report['year'][f'{int(quarter)}-quarter'] = {'total': quarters_sum.get(quarter)}
            for month in dataset.loc[dataset.quarter == quarter].month.unique():
                report['year'][f'{int(quarter)}-quarter'][f'{int(month)}-month'] = {'total': months_sum.get(month)}
                for day, day_value in dataset.loc[dataset.month == month][['day', 'hash']].values:
                    report['year'][f'{int(quarter)}-quarter'][f'{int(month)}-month'][f'{int(day)}-day'] = {
                        'total': day_value}

        return report


def month_day_report(db, year, month):
    statement = db.query(db_hashrates.Hashrate).filter(extract('year', db_hashrates.Hashrate.date) == year).filter(extract('month', db_hashrates.Hashrate.date) == month).statement
    dataset = pd.read_sql(statement, engine)
    dataset['date'] = pd.to_datetime(dataset.date, format='%Y-%m-%d')
    dataset['day']: Series = dataset.date.dt.day
    report = [dataset.hash.sum()]
    for day, hash, average in dataset[['day', 'hash', 'average']].values:
        report.append({'total': hash, 'average': average})

    return report


get_func_by_report_type = {
    'month_day': month_day_report
}


