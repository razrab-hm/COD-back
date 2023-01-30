from enum import Enum

import openpyxl
import pandas as pd
from fastapi import UploadFile
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame
from roman import toRoman


class Style(Enum):
    FONT = Font(bold=True)
    FILL = PatternFill("solid", fgColor="00C0C0C0")
    DATA_FILL = PatternFill("solid", fgColor="00CCFFCC")


def get_xls_data(file: UploadFile):
    xls = pd.ExcelFile(file.file)
    sheet: DataFrame = xls.parse(0)
    if type(sheet.values[0][0]) == str:
        return zip(sheet[sheet.keys()[0]], sheet[sheet.keys()[1]])
    else:
        return zip(sheet[sheet.keys()[1]], sheet[sheet.keys()[0]])


def initialize_workbook(titles):
    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40

    insert_data(ws, titles, 1, True)

    return wb, ws


def insert_data(ws, data, row_counter, is_header=False):
    if not is_header:
        for column, value in enumerate(data):
            cell = ws.cell(row=row_counter, column=column + 1, value=value)
            if row_counter % 2 == 1:
                cell.fill = Style.DATA_FILL.value
    else:
        for column, value in enumerate(data):
            cell = ws.cell(row=row_counter, column=column + 1, value=value)
            cell.font = Style.FONT.value
            cell.fill = Style.FILL.value


def month_day_report(dataset, year):
    wb, ws = initialize_workbook(['Date', 'Year', 'Days Hashrate (EH)'])
    row_counter = 2

    for day, hash_rate, average, month_name in dataset[['day', 'hash', 'average', 'month_name']].values:
        insert_data(ws, [f'{month_name} {day}, {year}', year, hash_rate], row_counter)
        row_counter += 1

    insert_data(ws, ['Totals:', year, dataset.hash.sum()], row_counter, True)

    wb.save("sample.xlsx")

    return {'total': dataset.hash.sum()}


def year_quarter_month_report(dataset, quarter_groups, months_sum, quarter_sum, year):
    wb, ws = initialize_workbook(['Month', 'Year', 'Month hashrate (EH)'])
    row_counter = 2

    for quarter in quarter_groups:
        for month_name in dataset.loc[dataset.quarter == quarter[0]].month_name.unique():
            insert_data(ws, [month_name, year, months_sum.get(month_name)], row_counter)
            row_counter += 1

        insert_data(ws, [f'{toRoman(quarter[0])} Quarter', year, quarter_sum.get(quarter[0])], row_counter, True)
        row_counter += 1

    insert_data(ws, ['Totals:', year, dataset.hash.sum()], row_counter, True)

    wb.save("sample.xlsx")

    return {'total': dataset.hash.sum()}


def year_quarter_report(dataset, quarters_sum, year):
    wb, ws = initialize_workbook(['Quarter', 'Year', 'Quarter hashrate (EH)'])
    row_counter = 2

    for quarter_pk, quarter_sum in quarters_sum.items():
        insert_data(ws, [f'{toRoman(quarter_pk)} Quarter', year, quarter_sum], row_counter)
        row_counter += 1

    insert_data(ws, ['Totals:', year, dataset.hash.sum()], row_counter, True)

    wb.save("sample.xlsx")

    return {'total': dataset.hash.sum()}


def year_quarter_month_day_report(dataset, quarter_groups, year, months_sum, quarter_sum, months_sum_average, quarter_sum_average):
    pass


def quarter_month_report(dataset, month_sums, month_names):
    pass


def quarter_month_day_report(dataset, months_sum):
    pass
